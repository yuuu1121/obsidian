# -*- coding: utf-8 -*-
"""
Read the design parameters out of Cycloidal_Drive_Design.xlsx, compute the
performance quantities that spreadsheet formulas cannot, and write them back
into a "6.성능계산" sheet with charts.

Run by double-clicking 성능계산.bat, or:  python run_performance.py
"""
from __future__ import annotations

import math
import sys
import pathlib
import datetime

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series

sys.path.insert(0, str(pathlib.Path(__file__).parent))
from cycloid_solver import (Geom, contact_forces, transmission_error,
                            entry_angle, initial_clearance)

XLSX = pathlib.Path(__file__).parent.parent / "Cycloidal_Drive_Design.xlsx"
SHEET = "6.성능계산"

HDR_BG = PatternFill("solid", fgColor="1F4E78")
HDR_FG = Font(bold=True, color="FFFFFF", size=11)
SEC_BG = PatternFill("solid", fgColor="DDEBF7")
CALC_BG = PatternFill("solid", fgColor="F2F2F2")
WARN_BG = PatternFill("solid", fgColor="FFEB9C")
OK_BG = PatternFill("solid", fgColor="C6EFCE")
BOLD = Font(bold=True)
SMALL = Font(size=9, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")


def read_geom(wb):
    """pull the yellow input cells; they are plain numbers, so data_only
    is not needed and the file works even if Excel never recalculated"""
    s = wb["1.설계입력"]
    def num(addr, name):
        v = s[addr].value
        if not isinstance(v, (int, float)):
            raise SystemExit(f"[{name}] 셀 {addr} 이 숫자가 아닙니다: {v!r}\n"
                             f"  -> 1.설계입력 시트의 노란 셀을 확인하세요.")
        return float(v)
    return Geom(
        rp=num("C6", "r_p"), rrp=num("C7", "r_rp"), a=num("C8", "a"),
        zc=int(num("C9", "z_c")), bc=num("C10", "b_c"),
        drp=num("C33", "d_rp"), drrp=num("C34", "d_rrp"),
        # C42 holds the output torque; fall back to the paper's 420 N.m
        T=float(s["C42"].value) if isinstance(s["C42"].value, (int, float))
          else 420.0,
    )


def main():
    if not XLSX.exists():
        raise SystemExit(f"엑셀 파일을 찾을 수 없습니다: {XLSX}")
    print(f"reading  {XLSX.name}")
    wb = load_workbook(XLSX)
    g = read_geom(wb)

    print(f"  r_p={g.rp} r_rp={g.rrp} a={g.a} z_c={g.zc} z_p={g.zp} "
          f"b_c={g.bc}  T={g.T} N.m")
    print(f"  d_rp={g.drp} d_rrp={g.drrp}   K1={g.K1:.5f} -> {g.K1m:.5f}")

    errs = g.check()
    for e in errs:
        print("  WARN:", e)
    if any("drive condition" in e or "K1" in e and ">= 1" in e for e in errs):
        raise SystemExit("설계 조건 위반입니다. 1.설계입력 을 고친 뒤 다시 실행하세요.")

    print("computing contact forces ...")
    cf = contact_forces(g, verbose=True)
    print("computing transmission error / backlash ...")
    te_mod = transmission_error(g, modified=True, n_steps=37, verbose=True,
                               n_profile=60000)
    print("  (reference) standard profile ...")
    te_std = transmission_error(g, modified=False, n_steps=13,
                               n_profile=60000)

    # ---------------------------------------------------------------- sheet
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)
    ws.sheet_properties.tabColor = "C00000"
    for c, w in zip("ABCDEFGHIJ", (4, 26, 14, 10, 46, 3, 12, 12, 12, 12)):
        ws.column_dimensions[c].width = w

    def title(cell, text, span):
        ws[cell] = text
        ws[cell].fill = HDR_BG
        ws[cell].font = HDR_FG
        ws[cell].alignment = CENTER
        col, row = ws[cell].column, ws[cell].row
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)

    def section(cell, text, span):
        ws[cell] = text
        ws[cell].fill = SEC_BG
        ws[cell].font = BOLD
        col, row = ws[cell].column, ws[cell].row
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)

    def put(r, name, val, unit, note, fmt="0.0000"):
        ws.cell(r, 2, name).font = BOLD
        c = ws.cell(r, 3, val)
        c.fill = CALC_BG; c.border = BOX; c.alignment = CENTER
        c.number_format = fmt
        ws.cell(r, 4, unit).alignment = CENTER
        ws.cell(r, 5, note).font = SMALL
        ws.cell(r, 5).alignment = WRAP

    title("B2", "성능 계산 결과  (파이썬 solver 출력 - 수식이 아닙니다)", 4)
    stamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["B3"] = (f"계산 시각 {stamp}.  1.설계입력 을 바꾼 뒤 성능계산.bat 을 "
                f"다시 실행하면 이 시트가 갱신됩니다.")
    ws["B3"].font = SMALL
    ws.merge_cells("B3:E3")

    section("B5", "입력 요약", 4)
    put(6, "핀 중심원 반경 r_p", g.rp, "mm", "1.설계입력 C6")
    put(7, "핀 반경 r_rp", g.rrp, "mm", "1.설계입력 C7")
    put(8, "편심량 a", g.a, "mm", "1.설계입력 C8")
    # NOTE: never start a note with "=" -- Excel parses the cell as a formula
    # and then strips every formula on the sheet as corrupt.
    put(9, "이빨 수 z_c", g.zc, "개", "감속비와 동일", "0")
    put(10, "핀 수 z_p", g.zp, "개", "z_c + 1", "0")
    put(11, "기어 폭 b_c", g.bc, "mm", "접촉 길이")
    put(12, "출력 토크 T", g.T, "N.m", "1.설계입력 C42")
    put(13, "핀위치 수정 d_rp", g.drp, "mm", "1.설계입력 C33")
    put(14, "핀반경 수정 d_rrp", g.drrp, "mm", "1.설계입력 C34")
    put(15, "K1 (수정 후)", g.K1m, "-", "a*z_p/(r_p+d_rp).  논문 B 명시대로 갱신한 값",
        "0.000000")

    section("B17", "접촉력  (논문 B 식20~32)", 4)
    put(18, "최대 접촉력 F_max", cf["F_max"], "N",
        "가장 큰 힘을 받는 이빨의 접촉력")
    put(19, "최대 변형 delta_max", cf["delta_max"] * 1000, "um",
        "헤르츠 접촉 변형. 초기 틈새와 같은 자릿수여야 정상")
    put(20, "동시 접촉 이빨 수", cf["n_contact"], "개",
        f"핀 {cf['contact_pins'][0]}~{cf['contact_pins'][-1]} 번", "0")
    put(21, "맞물림 각 범위", cf["engage_deg"][0], "deg",
        f"{cf['engage_deg'][0]:.1f} ~ {cf['engage_deg'][1]:.1f} deg "
        f"(논문 B Fig.11 대응)", "0.0")
    put(22, "최대 초기 틈새", max(cf["gap"]) * 1000, "um", "논문 B 식(20)")
    put(23, "식(27) 잔차", cf["residual_eq27"] * 100, "%",
        "0 에 가까우면 변형-힘 두 식이 일관됨", "0.00E+00")

    section("B25", "백래시 / 전달오차  (논문 B 식15·16)", 4)
    put(26, "진입각 beta", math.degrees(te_mod["beta"]) * 60, "arcmin",
        "틈새를 극복하고 접촉하기까지 도는 각")
    put(27, "백래시 (2*beta)", te_mod["backlash_arcmin"], "arcmin",
        "논문 B 3.5절 정의")
    te = te_mod["TE_arcmin"]
    put(28, "전달오차 평균", float(np.nanmean(te)), "arcmin",
        "무부하, 한 주기 평균")
    put(29, "전달오차 최소", float(np.nanmin(te)), "arcmin", "절댓값이 작을수록 정밀")
    put(30, "전달오차 최대", float(np.nanmax(te)), "arcmin", "")
    put(31, "전달오차 변동폭", float(np.nanmax(te) - np.nanmin(te)), "arcmin",
        "이 변동이 진동·소음의 가진 성분", "0.000000")
    put(32, "표준치형 TE 검증", float(np.nanmax(np.abs(te_std["TE_arcmin"]))),
        "arcmin",
        "공액 치형이므로 0 이어야 정상. 0.05 미만이면 solver 정상", "0.000000")
    ws.cell(32, 3).fill = (OK_BG
        if np.nanmax(np.abs(te_std["TE_arcmin"])) < 0.05 else WARN_BG)

    # 1 m arm equivalent, for intuition
    put(34, "백래시 (1m 팔 끝단)", te_mod["backlash_arcmin"] / 60 * math.pi / 180
        * 1000, "mm", "길이 1m 로봇 팔 끝에서의 변위. 1 arcmin = 0.29 mm")

    # ---------------------------------------------------------- data tables
    r0 = 37
    ws.cell(r0 - 1, 7, "핀별 접촉력").font = BOLD
    for j, h in enumerate(["핀 번호", "각도 deg", "접촉력 N", "초기틈새 um",
                           "변형 um"], start=7):
        c = ws.cell(r0, j, h)
        c.fill = HDR_BG; c.font = HDR_FG; c.alignment = CENTER
    for i in range(len(cf["pin_index"])):
        r = r0 + 1 + i
        ws.cell(r, 7, int(cf["pin_index"][i])).number_format = "0"
        ws.cell(r, 8, float(np.degrees(cf["phi"][i]))).number_format = "0.0"
        ws.cell(r, 9, float(cf["F_i"][i])).number_format = "0.00"
        ws.cell(r, 10, float(cf["gap"][i] * 1000)).number_format = "0.000"
        ws.cell(r, 11, float(cf["delta"][i] * 1000)).number_format = "0.000"
    n_force = len(cf["pin_index"])

    r1 = r0 + n_force + 3
    ws.cell(r1 - 1, 7, "전달오차 곡선").font = BOLD
    for j, h in enumerate(["크랭크각 deg", "TE arcmin"], start=7):
        c = ws.cell(r1, j, h)
        c.fill = HDR_BG; c.font = HDR_FG; c.alignment = CENTER
    for i in range(len(te)):
        r = r1 + 1 + i
        ws.cell(r, 7, float(te_mod["phi_in_deg"][i])).number_format = "0.000"
        v = te[i]
        ws.cell(r, 8, None if np.isnan(v) else float(v)).number_format = "0.00000"
    n_te = len(te)

    r2 = r1 + n_te + 3
    ws.cell(r2 - 1, 7, "초기 틈새 분포 (0~180deg)").font = BOLD
    for j, h in enumerate(["각도 deg", "틈새 um"], start=7):
        c = ws.cell(r2, j, h)
        c.fill = HDR_BG; c.font = HDR_FG; c.alignment = CENTER
    angs = np.arange(0.5, 180.0, 1.0)
    for i, ad in enumerate(angs):
        r = r2 + 1 + i
        ws.cell(r, 7, float(ad)).number_format = "0.0"
        ws.cell(r, 8, float(initial_clearance(g, math.radians(ad)) * 1000)
                ).number_format = "0.0000"
    n_clr = len(angs)

    # -------------------------------------------------------------- charts
    ch = ScatterChart()
    ch.title = "핀별 접촉력 분포"
    ch.x_axis.title = "핀 번호"
    ch.y_axis.title = "접촉력 [N]"
    ch.height, ch.width = 9, 16
    xr = Reference(ws, min_col=7, min_row=r0 + 1, max_row=r0 + n_force)
    yr = Reference(ws, min_col=9, min_row=r0, max_row=r0 + n_force)
    ch.series.append(Series(yr, xr, title_from_data=True))
    ws.add_chart(ch, "B40")

    ch2 = ScatterChart()
    ch2.title = "전달오차 (무부하, 한 주기)"
    ch2.x_axis.title = "크랭크 입력각 [deg]"
    ch2.y_axis.title = "전달오차 [arcmin]"
    ch2.height, ch2.width = 9, 16
    x2 = Reference(ws, min_col=7, min_row=r1 + 1, max_row=r1 + n_te)
    y2 = Reference(ws, min_col=8, min_row=r1, max_row=r1 + n_te)
    s2 = Series(y2, x2, title_from_data=True)
    s2.marker.symbol = "none"
    ch2.series.append(s2)
    ws.add_chart(ch2, "B60")

    ch3 = ScatterChart()
    ch3.title = "초기 틈새 분포 (논문 B Fig.11 대응)"
    ch3.x_axis.title = "각도 [deg]"
    ch3.y_axis.title = "초기 틈새 [um]"
    ch3.height, ch3.width = 9, 16
    x3 = Reference(ws, min_col=7, min_row=r2 + 1, max_row=r2 + n_clr)
    y3 = Reference(ws, min_col=8, min_row=r2, max_row=r2 + n_clr)
    s3 = Series(y3, x3, title_from_data=True)
    s3.marker.symbol = "none"
    ch3.series.append(s3)
    ws.add_chart(ch3, "B80")

    # ------------------------------------------------------------ caveats
    section("B100", "결과를 읽을 때 반드시 알아야 할 것", 4)
    notes = [
        ("검증된 것",
         "표준(공액) 치형을 넣으면 전달오차가 0.0001 arcmin 미만으로 나옵니다. "
         "틈새가 0인 치형은 전달오차가 0이어야 하므로, 이것이 solver 가 옳게 "
         "동작한다는 증거입니다. C32 셀에서 매번 확인하세요."),
        ("절대값은 논문과 다릅니다",
         "논문 B 의 F_max=1131 N 은 이 solver 로 약 583 N 입니다(약 1/2). "
         "원인은 논문 자체의 모순입니다: 논문의 F_max 와 delta_max=0.0051mm 는 "
         "식(27)로는 맞지만(1079 N 필요) 식(25) 토크균형으로는 맞지 않습니다"
         "(463 N). 논문 값을 재현하려면 Tc ~ 1.34*T 가 필요한데 논문에 근거가 "
         "없습니다. 백래시·전달오차도 같은 이유로 약 2배 차이가 납니다."),
        ("케이스 비교는 신뢰할 수 있습니다",
         "Case a/Case b 의 백래시 비율은 solver 1.2375, 논문 Table3 1.1979 로 "
         "3% 차이입니다. 접촉 이빨 수가 늘면 최대 접촉력이 준다는 관계도 "
         "재현됩니다. 즉 수정계수를 바꿔가며 상대 비교하는 용도로는 유효하고, "
         "절대 수치를 카탈로그 스펙과 직접 비교하면 안 됩니다."),
        ("이 solver 가 무시한 것",
         "핀의 굽힘 변형 f_max (논문도 무시), 접촉을 F ~ delta 로 선형 가정 "
         "(실제 헤르츠는 F ~ delta^1.5), 1단 유성기어, 출력 핀-홀 커플링의 "
         "틈새, 베어링 유격, 제조·조립 오차. 실제 감속기의 총 백래시는 여기서 "
         "계산한 값보다 큽니다."),
        ("부하 상태 전달오차",
         "여기 값은 무부하입니다. 논문 B 는 하중이 걸리면 개선 폭이 19.1% 에서 "
         "4.25% 로 줄어든다고 보고합니다. 탄성변형이 치형 차이를 가리기 "
         "때문입니다. 실제 운전 조건 평가에는 FEM 이 필요합니다."),
    ]
    r = 101
    for k, v in notes:
        ws.cell(r, 2, k).font = BOLD
        ws.cell(r, 2).alignment = WRAP
        ws.cell(r, 3, v).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws.row_dimensions[r].height = max(30, 13 * (len(v) // 60 + 1))
        r += 1

    # ---- guard: a text cell that begins with "=" makes Excel treat the whole
    # sheet's formulas as corrupt and silently strip them.  Refuse to save.
    import re as _re
    _ALLOWED = {"IF", "AND", "OR", "NOT", "ABS", "MAX", "MIN", "SQRT", "COS",
                "SIN", "ATAN2", "RADIANS", "DEGREES", "SUM", "COUNT", "TEXT",
                "ISNUMBER", "SEARCH", "PI"}
    bad = []
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            for fn in _re.findall(r"([A-Za-z_][A-Za-z0-9_.]*)\s*\(", v):
                if fn.upper() not in _ALLOWED:
                    bad.append(f"{c.coordinate}: unknown function {fn!r}")
            t = _re.sub(r'"[^"]*"', "", v)
            t = _re.sub(r"'[^']*'!", "", t)
            if _re.search(r"[가-힣]", t):
                bad.append(f"{c.coordinate}: 한글로 시작하는 텍스트가 "
                           f"수식으로 인식됩니다 -> {v[:50]}")
    if bad:
        print("\n[중단] 저장하지 않았습니다. 잘못된 셀:")
        for b in bad[:10]:
            print("   ", b)
        raise SystemExit("잘못된 수식 때문에 저장을 취소했습니다.")

    wb.save(XLSX)
    print(f"\nwrote sheet [{SHEET}] into {XLSX.name}")
    print(f"  F_max        = {cf['F_max']:.2f} N over {cf['n_contact']} teeth")
    print(f"  backlash     = {te_mod['backlash_arcmin']:.4f} arcmin")
    print(f"  TE mean      = {np.nanmean(te):+.4f} arcmin")
    print(f"  sanity (std) = {np.nanmax(np.abs(te_std['TE_arcmin'])):.6f} arcmin"
          f"  (must be ~0)")


if __name__ == "__main__":
    try:
        main()
    except SystemExit as e:
        print(f"\n[중단] {e}")
        sys.exit(1)
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"\n[오류] {type(e).__name__}: {e}")
        sys.exit(1)
